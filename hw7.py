'''
First, insure that you have installed openpyxl. Remember, you install it by
running to following from a command prompt/terminal/shell:

(WINDOWS)
  pip install openpyxl
(MAC/LINUX/MIMIR)
  pip3 install openpyxl

You may already have the latest version, so the above may not be needed.

Running is the same as always:

(WINDOWS)
  python hw7.py
(MAC/LINUX/MIMIR)
  python3 hw7.py

Complete ALL of the following functions (removing the "pass" and replacing
it with real code). If you need any additional imports, add them below.
'''
import openpyxl as xl
import os
import datetime

def grail():
    '''
    Query the user for:

    1) Their Name ("What is your name? ")
    2) Their Quest ("What is your quest? ")
    3) Their Favorite Color ("What is your favorite color? ")

    Next:

    1) Open a "three_questions.txt" file so that lines of text can be APPENDED.
    Use a with/as block to avoid needing to manually close the file.
    2) If they enter "to seek the holy grail" (case insensitive) as their quest
    AND if they enter "blue" (case insensitive) as their favorite color, then
    write a line to the file that contains a timestamp (see below) followed by
    the words "{name} was allowed to cross the Bridge of Death." Don't forget 
    the newline at the end.
    3) Otherwise, write a line to the file that contains a timestamp followed by
    the words "{name} met their maker in a bubbling pool of lava."

    A timestamp is just a string containing the current date and time. An easy
    way to get this in Python is using the datetime object and calling the now
    method. Speaking of now, now would ge a good time to dust off your Google
    skills and figure out how to do that. Or perhaps open the textbook to that
    chapter.

    When all is done (and assuming this function is called a couple times), the
    contents of three_questions.txt might look similar to:

    2019-11-13 17:27:43.163137: Paul met their maker in a bubbling pool of lava.
    2019-11-13 17:30:15.610515: Paul was allowed to cross the Bridge of Death.
    '''
    name = input("What is your name? ")
    quest = input("What is your quest? ")
    color = input("What is your favorite color? ")
    name = name.strip()
    quest = quest.strip()
    color = color.strip()
    now = datetime.datetime.now()
    with open("three_questons.txt", "w") as file:
        holyGrail = 'to seek the holy grail'
        if quest.upper() == holyGrail.upper() and color.upper() == "BLUE":
            file.write(f"{now}: {name} was allowed to cross the Bridge of Death.\n")
        else:
            file.write(f"{now}: {name} met their maker in a bubbling pool of lava.\n")

def file_math(filename):
    '''
    This function should perform a specified mathematical operation on a 
    sequence of integers stored in a file. The file will contain multiple lines.
    Each line holds the operation and the integers all separated by commas.
    For example:

    +,5,11
    /,25,5
    +,6,5,4,3

    Your job is to read the file and print out the results. The output for a
    file containing the above would be:

    16
    5.0
    18

    Note that the 5.0 is because Python always returns the result of division
    as a float. The steps are basically as follows:

    1) Open the file (again the filename is passed in...no need for you to ask
    for it) in read only mode. Use a with/as here to avoid needing to close
    the file
    2) Iterate over the file line by line
        a) "split" a line by the comma
        b) Use an if/elif/else block to perform code specific to each of the four
        supported operations ('+', '-', '*' and '/')
        c) Use an accumulator variable (aka a subtotal variable) and a for loop
        to add or subtract or multiply or divide each of the integers in
        sequence.
        d) Print out the result

    Since I'm nice, here's some example code to do 2c & 2d. It's assuming that
    the operator is '+' and that the results of the split operation are stored
    in a variable called parts:

        nums = parts[1:]            # everything except for the operator
        subtotal = int(nums[0])     # start with the first num

        for num in nums[1:]:        # loop over the remaining nums
            subtotal += int(num)    # add the number to the subtotal
            
        print(subtotal)             # print the final total

    mathtest.txt contains a sample file that you can use to test your function.
    '''
    with open(filename, "r") as file:
        for line in file:
            line = line.strip()
            parts = []
            parts = line.split(",")
            sign = parts[0]         # operator
            nums = parts[1:]        # numbers
            if sign == '+':
                subtotal = int(nums[0])
                for num in nums[1:]:
                    subtotal += int(num)
            elif sign == '-':
                subtotal = int(nums[0])
                for num in nums[1:]:
                    subtotal -= int(num)   
            elif sign == '*':
                subtotal = int(nums[0])
                for num in nums[1:]:
                    subtotal *= int(num)   
            else:
                subtotal = int(nums[0])
                for num in nums[1:]:
                    subtotal /= int(num)                                          
            print(subtotal)
                    
def simple_calendar():
    '''
    The goal is to create a very simple calendar in a Microsoft Excel
    spreadsheet.

    1) Query the user for the name of the month (do basic error checking
    here to insure it is actually one of the, you know, months).
    2) Create a new openpyxl Workbook.
    3) Use the "active" Worksheet (i.e., no need to create a new worksheet)
    4) In Cell D1 write the name of the month
    5) In Cells A2 through G2, write the names of the weekdays
    6) In Cells A3 through G7 (well, the rectanglar box bounded by these cells),
    write the numbers 1 through 31 such that 1 is in A3 and 31 is in C7. Use
    for in range loops to do this.
    7) Save the workbook as 'calendar.xlsx'.

    That's all you have to do. But for fun, you can:

    a) format it nice and pretty-like.
    b) make the calendar ACTUALLY match the month for the current year by
    putting the day numbers into the proper column for day of the week...
    this may help: https://www.mytecbits.com/internet/python/day-of-week
    '''
    month = input("Enter the name of the month: ")
    month = month.strip()
    months = ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december']
    weekdays = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    dayofweek = 0
    date = 1
    if month.lower() in months:
        wb = xl.Workbook()
        ws = wb.active
        ws['D1'] = month
        for col in ws['A2:G2']:                     # Setting the Day of the Week
            for cell in col:
                cell.value = weekdays[dayofweek]
                dayofweek += 1
        for row in ws['A3:G7']:                     # Setting the dates
            for cell in row:
                cell.value = date
                date += 1
        wb.save('calendar.xlsx')
    else:
        print("Invalid month entered. Please try again.")

def excel_math(filename):
    '''
    This is exactly the same thing as file math. Except this time the operations
    and numbers are stored in an excel spreadsheet. So instead of opening a
    text file and processing line by line, you'll open a spreadsheet and
    process it row by row. Same deal, though. The first cell in every row has
    a +, -, * or / character. The rest will have integers.

    mathtest.xlsx contains a sample file that you can use to test your function.
    '''
    wb = xl.open(filename)
    ws = wb.active

    for stuff in ws.values:
        sign = stuff[0]
        nums = stuff[1:]
        #print(sign)
        #print(nums)
        if sign == '+':
            subtotal = int(nums[0])
            for num in nums[1:]:
                if num == None:
                    break
                subtotal += int(num)
        elif sign == '-':
            subtotal = int(nums[0])
            for num in nums[1:]:
                if num == None:
                    break                
                subtotal -= int(num)   
        elif sign == '*':
            subtotal = int(nums[0])
            for num in nums[1:]:
                if num == None:
                    break                
                subtotal *= int(num)   
        else:
            subtotal = int(nums[0])
            for num in nums[1:]:
                if num == None:
                    break                
                subtotal /= int(num)                                          
        print(subtotal)        

# THIS PART IS MY CODE. IT CALLS YOUR CODE. NO NEED TO THOUCH IT.

while True:
    print('MAIN MENU'.center(40))
    print('=' * 40)
    print('1) (Monty) Python Game')
    print('2) File Math')
    print('3) Simple Calendar')
    print('4) Excel Math')
    print('5) Exit')
    print('=' * 40)

    choice = input('Choice: ')
    if choice == '1':
        grail()
    elif choice == '2':
        the_name = input("Enter filename: ")
        if not the_name.endswith('.txt') or not os.path.isfile(the_name):
            print("The file you entered is invalid.")
            continue
        file_math(the_name)
    elif choice == '3':
        simple_calendar()
    elif choice == '4':
        the_name = input("Enter filename: ")
        if not the_name.endswith('.xlsx') or not os.path.isfile(the_name):
            print("The file you entered is invalid.")
            continue
        excel_math(the_name)
    elif choice == '5':
        print('Goodbye!')
        break
    else:
        print('INVALID SELECTION!!')
        print()

# when the script gets here, it exits...all done!